from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os

# Define the specific path
folder_path = r"C:\Users\JB VILLE\OneDrive\Desktop\fafa\PY\EXCEL\excel file"  # Update this path
file_name = "recreated_class_schedule2.xlsx"
file_path = os.path.join(folder_path, file_name)

# Ensure the folder exists
os.makedirs(folder_path, exist_ok=True)

# Create workbook and worksheet
wb = Workbook()
ws = wb.active
ws.title = "JB Class Schedule"

# Define the schedule headers and content
days = ["Time", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
schedule = [
    ["7:30 AM - 9:30 AM", "", "", "", "", "", "Operating System (LEC)", "Intro to Web Design (LEC)"],
    ["9:30 AM - 12:30 PM", "NSTP 2", "", "Fundamentals of Information Management", "", "", "Operating System (LAB)", "Intro to Web Design (LAB)"],
    ["1:00 PM - 2:30 PM", "", "", "", "", "Computer Programming 2 (LAB)", "", ""],
    ["1:30 PM - 3:30 PM", "", "Computer Programming 2 (LEC)", "", "", "", "", ""],
    ["2:00 PM - 4:00 PM", "", "", "", "", "", "PE-2 Fitness Exercises", ""],
    ["9:00 AM - 10:30 AM", "", "", "", "The Contemporary World", "", "", ""],
    ["10:30 AM - 12:00 PM", "", "", "", "Art Appreciation", "", "", ""],
]

# Add headers to the worksheet with styles
header_fill = PatternFill(start_color="6DCDE8", end_color="6DCDE8", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

for col, day in enumerate(days, start=1):
    cell = ws.cell(row=1, column=col, value=day)
    cell.font = Font(bold=True, color="000000", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = header_fill
    cell.border = thin_border

# Add schedule rows with styles
for row_idx, row_data in enumerate(schedule, start=2):
    for col_idx, cell_data in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=cell_data)
        cell.font = Font(size=10, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

# Adjust column widths
column_widths = [20, 25, 25, 25, 25, 25, 30, 30]
for col, width in enumerate(column_widths, start=1):
    ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

# Adjust row heights
row_heights = [25] + [20] * len(schedule)
for row_idx, height in enumerate(row_heights, start=1):
    ws.row_dimensions[row_idx].height = height

# Save the workbook to the specified path
wb.save(file_path)

print(f"File saved successfully at: {file_path}")
