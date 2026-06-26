from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os

# === PATH SETTINGS ===
folder_path = r"PY\EXCEL"
file_name = "class_schedule_final1.xlsx"
file_path = os.path.join(folder_path, file_name)
os.makedirs(folder_path, exist_ok=True)

# === CREATE WORKBOOK ===
wb = Workbook()
ws = wb.active
ws.title = "Class Schedule"

# === HEADERS ===
days = ["Time", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

# === SCHOOL CLASS SCHEDULE ===
schedule = [
    ["07:30 AM - 09:30 AM", "", "", "", "", "System Integration & Architecture 1 LEC", "Applications Dev & Emerging Tech LEC", ""],
    ["08:00 AM - 3:00 PM", "", "", "", "", "", "", "Community Development"],
    ["09:30 AM - 12:30 PM", "", "", "", "", "System Integration & Architecture 1 LAB", "Applications Dev & Emerging Tech LAB", ""],
    ["01:00 PM - 04:00 PM", "Living in the IT Era", "Gender and Society", "", "", "", "", ""],
    ["01:30 PM - 03:30 PM", "", "", "", "", "Prof. Course Specialization 1 LEC", "Prof. Course Elective 1 LEC", ""],
    ["03:30 PM - 06:30 PM", "", "", "", "", "Prof. Course Specialization 1 LAB", "Prof. Course Elective 1 LAB", ""],
]

# Keep the Time column and only include days that have at least one subject.
included_columns = [0]
for col_idx in range(1, len(days)):
    if any(len(row) > col_idx and str(row[col_idx]).strip() for row in schedule):
        included_columns.append(col_idx)

days = [days[col_idx] for col_idx in included_columns]
schedule = [[row[col_idx] if len(row) > col_idx else "" for col_idx in included_columns] for row in schedule]

# === STYLES ===
header_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
time_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))
header_font = Font(bold=True, color="000000", size=12)
body_font = Font(size=11, color="000000")
time_font = Font(bold=True, size=11)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

# === HEADERS ===
for col, day in enumerate(days, start=1):
    cell = ws.cell(row=1, column=col, value=day)
    cell.font = header_font
    cell.alignment = center_align
    cell.fill = header_fill
    cell.border = thin_border

# === DATA ROWS ===
for row_idx, row_data in enumerate(schedule, start=2):
    for col_idx, cell_data in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=cell_data)
        cell.alignment = center_align
        cell.border = thin_border
        if col_idx == 1:  # Time column
            cell.fill = time_fill
            cell.font = time_font
        else:
            cell.font = body_font

# === COLUMN WIDTHS ===
all_column_widths = [18, 23, 23, 23, 23, 23, 24, 24]
column_widths = [all_column_widths[col_idx] for col_idx in included_columns]
for col, width in enumerate(column_widths, start=1):
    ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

# === ROW HEIGHTS ===
row_heights = [50] + [65] * len(schedule)
for row_idx, height in enumerate(row_heights, start=1):
    ws.row_dimensions[row_idx].height = height

# === SAVE FILE ===
try:
    wb.save(file_path)
    print(f"✅ Schedule saved successfully at: {file_path}")
except Exception as e:
    print(f"❌ Error while saving file: {e}")
